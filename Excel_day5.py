# import openpyxl module 
import openpyxl 
# Give the location of the file 
path = "studentdata.xlsx"
# To open the workbook 
# workbook object is created 
wb_obj = openpyxl.load_workbook(path) 
# Get workbook active sheet object 
# from the active attribute 
sheet_obj = wb_obj.active 
# Cell objects also have a row, column, 
# and coordinate attributes that provide 
# location information for the cell. 

# Cell object is created by using 
# sheet object's cell() method. 
cell_obj = sheet_obj.cell(row = 1, column = 1) 
# Print value of cell object 
# using the value attribute 
print(cell_obj.value) 



####################################

path = "studentdata.xlsx"
wb_obj = openpyxl.load_workbook(path)  
sheet_obj = wb_obj.active 
# Getting the value of maximum rows and column
row = sheet_obj.max_row
column = sheet_obj.max_column
print("Total Rows:", row)
print("Total Columns:", column)
print("\nValue of first column")
for i in range(1, row + 1): 
    cell_obj = sheet_obj.cell(row = i, column = 1) 
    print(cell_obj.value) 
print("\nValue of first row")
for i in range(1, column + 1): 
    cell_obj = sheet_obj.cell(row = 2, column = i) 
    print(cell_obj.value, end = " ")

############## reading excel ##############

path = "studentdata.xlsx"
wb_obj = openpyxl.load_workbook(path)  
sheet_obj = wb_obj.active 
# Cell object is created by using sheet object's cell() method. 
cell_obj = sheet_obj['A1': 'B6']
# Print value of cell object using the value attribute 
for cell1, cell2 in cell_obj:
    print(cell1.value, cell2.value)

########### writing

wb = openpyxl.Workbook() 
sheet = wb.active 
c1 = sheet.cell(row = 1, column = 1) 
# writing values to cells 
c1.value = "Hello"
c2 = sheet.cell(row= 1 , column = 2) 
c2.value = "World"
c3 = sheet['A2'] 
c3.value = "Welcome"
c4 = sheet['B2'] 
c4.value = "Everyone"
wb.save("sample.xlsx") 

###### append method ##########
 
wb = openpyxl.load_workbook("sample.xlsx")   
sheet = wb.active 
c = sheet['A3'] 
c.value = "New Data"
wb.save("sample.xlsx")

############## append #######
wb = openpyxl.load_workbook("sample.xlsx")   
sheet = wb.active   
data = (
    (1, 2, 3),
    (4, 5, 6))
for row in data:
    sheet.append(row)
wb.save('sample.xlsx')

########## arithmatic operation #####
wb = openpyxl.Workbook() 
sheet = wb.active 
# writing to the cell of an excel sheet 
sheet['A1'] = 200
sheet['A2'] = 300
sheet['A3'] = 400
sheet['A4'] = 500
sheet['A5'] = 600
# The value in cell A7 is set to a formula 
# that sums the values in A1, A2, A3, A4, A5 . 
sheet['A7'] = '= SUM(A1:A5)'
# save the file 
wb.save("sum.xlsx") 

################################
#Adjusting rows and merging cells

wb = openpyxl.Workbook() 
sheet = wb.active 
  
# writing to the specified cell 
sheet.cell(row = 1, column = 1).value = ' hello '
sheet.cell(row = 2, column = 2).value = ' everyone '
  
# set the height of the row 
sheet.row_dimensions[1].height = 70
  
# set the width of the column 
sheet.column_dimensions['B'].width = 20
  
# save the file 
wb.save('sample.xlsx')


############## merginf columns
wb = openpyxl.Workbook() 
sheet = wb.active 
  
# merge cell from A2 to D4 i.e. 
# A2, B2, C2, D2, A3, B3, C3, D3, A4, B4, C4 and D4 . 
# A2:D4' merges 12 cells into a single cell. 
sheet.merge_cells('A2:D4') 
sheet.cell(row = 2, column = 1).value = 'Twelve cells join together.'
  
# merge cell C6 and D6 
sheet.merge_cells('C6:D6')  
sheet.cell(row = 6, column = 6).value = 'Two merge cells.'
  
wb.save('sample.xlsx')

############### column 
wb = openpyxl.load_workbook('sample.xlsx') 
sheet = wb.active 
  
# unmerge the cells 
sheet.unmerge_cells('A2:D4')  
sheet.unmerge_cells('C6:D6')  
wb.save('sample.xlsx')

############ font style
# import Font function from openpyxl 
from openpyxl.styles import Font 
 
wb = openpyxl.Workbook() 
sheet = wb.active 
  
sheet.cell(row = 1, column = 1).value = "Welcome Python"
  
# set the size of the cell to 24 
sheet.cell(row = 1, column = 1).font = Font(size = 24 ) 
sheet.cell(row = 2, column = 2).value = "Welcome Python"
  
# set the font style to italic 
sheet.cell(row = 2, column = 2).font = Font(size = 24, italic = True) 
  
sheet.cell(row = 3, column = 3).value = "Welcome Python"
  
# set the font style to bold 
sheet.cell(row = 3, column = 3).font = Font(size = 24, bold = True) 
  
sheet.cell(row = 4, column = 4).value = "Welcome Python"
  
# set the font name to 'Times New Roman' 
sheet.cell(row = 4, column = 4).font = Font(size = 24, name = 'Times New Roman') 
  
wb.save('sample.xlsx')

######### chart 

# import BarChart class from openpyxl.chart sub_module
from openpyxl.chart import BarChart, Reference
# wb = openpyxl.Workbook()
wb = openpyxl.load_workbook("studentdata.xlsx")
sheet = wb.active
# write o to 9 in 1st column of the active sheet
for i in range(10):
    sheet.append([i])
    
# create data for plotting
values = Reference(sheet, min_col=1, min_row=1,
                   max_col=1, max_row=10)
chart = BarChart() # Create object of BarChart class
chart.add_data(values) #adding data to the Bar chart object
chart.title = " BAR-CHART "# set the title of the chart
chart.x_axis.title = " X_AXIS " # set the title of the x-axis
chart.y_axis.title = " Y_AXIS " # set the title of the y-axis
# add chart to the sheet the top-left corner of a chart is anchored to cell E2 .
sheet.add_chart(chart, "E2")
# save the file
wb.save("sample.xlsx")


###### working on images

# from openpyxl.drawing.image import Image
# wb = openpyxl.Workbook()  
# sheet = wb.active
# # Adding a row of data to the worksheet (used to 
# # distinguish previous excel data from the image) 
# sheet.append([10, 2010, "Flower", 4, "life"]) 
# # A wrapper over PIL.Image, used to provide image 
# # inclusion properties to openpyxl library 
# img = Image("flower.jpg")
# # Adding the image to the worksheet 
# # (with attributes like position) 
# sheet.add_image(img, 'A2') 
# # Saving the workbook created
# wb.save('sample.xlsx')


#######################################
from openpyxl.drawing.image import Image
  
wb = openpyxl.load_workbook("studentdata.xlsx")  
# Getting the sheetnames as a list using the sheetnames attribute
sheetNames=wb.sheetnames # Traversing in the sheetNames list
for name in sheetNames:
       print(name)


# #access worksheets
# sheet=wb.worksheets[1] #access employee worksheet
# cell_obj=sheet['A1':'B4']
# for c1,c2 in cell_obj:
#        print(c1.value, c2.value)
#access worksheets
sheet=wb.worksheets[0] #access student worksheet
cell_obj=sheet['A1':'B4']
for c1,c2 in cell_obj:
       print(c1.value, c2.value)





